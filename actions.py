import csv
import io
import requests

REFUND_HEADERS = [
    "Name of Contact",
    "Payout Link Amount",
    "Contact Phone Number",
    "Contact Email ID",
    "Send Link to Phone Number",
    "Send Link to Mail ID",
    "Contact Type",
    "Payout Purpose",
    "Payout Description",
    "Reference ID(optional)",
    "Internal notes(optional): Title",
    "Internal notes(optional): Description",
]


def generate_refund_csv(cases: list) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(REFUND_HEADERS)
    for c in cases:
        amount = c.get("extra_amount") or 0
        w.writerow([
            "Wiom User",
            int(amount) if amount else 0,
            c.get("customer_mobile", ""),
            "",
            "Yes",
            "No",
            "customer",
            "refund",
            "Refund for Extra money taken",
            c.get("ticket_id", ""),
            "",
            "",
        ])
    return out.getvalue()


def _aggregate_penalty(cases: list) -> dict:
    """Aggregate by partner, sum amounts (negative)."""
    partners: dict = {}
    for c in cases:
        pid = c.get("current_partner_account_id", "unknown")
        amt = c.get("extra_amount") or 0
        if pid not in partners:
            partners[pid] = {"amount": 0, "name": c.get("current_partner_name", "")}
        partners[pid]["amount"] += amt
    return partners


def generate_penalty_csv(cases: list) -> str:
    partners = _aggregate_penalty(cases)
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Partner Id", "Amount", "Remark"])
    for pid, data in partners.items():
        w.writerow([pid, -abs(data["amount"]), "Breach Fundamental Principle 2 (PX2026)"])
    return out.getvalue()


def generate_penalty_xlsx(cases: list) -> bytes:
    """Generate penalty Excel file with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    partners = _aggregate_penalty(cases)
    wb = Workbook()
    ws = wb.active
    ws.title = "Partner Penalty"

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["AccountId", "Amount", "Remark"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # Data rows
    row_idx = 2
    for pid, data in partners.items():
        # AccountId as integer (Sheets may return "12345.0")
        try:
            pid_int = int(float(pid))
        except (ValueError, TypeError):
            pid_int = pid
        ws.cell(row=row_idx, column=1, value=pid_int).border = thin_border
        amt_cell = ws.cell(row=row_idx, column=2, value=-abs(data["amount"]))
        amt_cell.number_format = '#,##0'
        amt_cell.border = thin_border
        ws.cell(row=row_idx, column=3, value="Breach Fundamental Principle 2 (PX2026)").border = thin_border
        row_idx += 1

    # Auto-width columns
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_b1_penalty_xlsx(cases: list) -> bytes:
    """Generate B1 (FP1) penalty Excel file. One row per case, -2000 each."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "FP1 Partner Penalty"

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["AccountId", "Amount", "Remark"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # Data rows — one row per case (NOT aggregated per partner)
    row_idx = 2
    for c in cases:
        pid = c.get("partner_id") or ""
        try:
            pid_int = int(float(pid))
        except (ValueError, TypeError):
            pid_int = pid
        ws.cell(row=row_idx, column=1, value=pid_int).border = thin_border
        amt_cell = ws.cell(row=row_idx, column=2, value=-2000)
        amt_cell.number_format = '#,##0'
        amt_cell.border = thin_border
        ws.cell(row=row_idx, column=3, value="Breach Fundamental Principle 1").border = thin_border
        row_idx += 1

    # Auto-width columns
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_partner_comms_csv(cases: list) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Partner Mobile", "Technician Name", "Amount"])
    for c in cases:
        w.writerow([
            c.get("partner_mobile", ""),
            c.get("technician_name") or c.get("install_name", ""),
            c.get("extra_amount", 0),
        ])
    return out.getvalue()


def send_to_slack(webhook_url: str, text: str, csv_string: str = None) -> dict:
    if not webhook_url:
        return {"ok": False, "error": "No Slack webhook URL configured"}

    msg_text = text
    if csv_string:
        lines = csv_string.strip().split("\n")
        preview = "\n".join(lines[:8])
        if len(csv_string) <= 2800:
            msg_text += f"\n```\n{csv_string}\n```"
        else:
            msg_text += f"\n```\n{preview}\n... ({len(lines) - 1} rows total)\n```"

    try:
        resp = requests.post(webhook_url, json={"text": msg_text}, timeout=15)
        if resp.status_code == 200:
            return {"ok": True}
        return {"ok": False, "error": resp.text}
    except Exception as e:
        return {"ok": False, "error": str(e)}

import json
import os
import queue
import threading
import functools
from flask import Flask, jsonify, request, render_template, Response, stream_with_context
from flask_cors import CORS

import db
import sheets_db
import sheets_db_b1 as b1db
import config
import actions

def _parse_expiry_date(val) -> str:
    """Convert Excel serial date or DD-MM-YYYY string to YYYY-MM-DD. Returns '' if unparseable."""
    if val is None or val == "":
        return ""
    from datetime import datetime, timedelta
    # Excel serial integer (e.g. 46019 → 2025-12-28)
    try:
        n = float(str(val).replace(",", ""))
        if 40000 < n < 60000:  # plausible date serial range (~2009–2064)
            return (datetime(1899, 12, 30) + timedelta(days=int(n))).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass
    # Text date DD-MM-YYYY or DD/MM/YYYY
    s = str(val).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s  # store as-is if nothing matched


def _normalize_partner_id(val) -> str:
    """Normalize partner IDs: strip .0, handle scientific notation."""
    if not val and val != 0:
        return ""
    try:
        return str(int(float(str(val))))
    except (ValueError, TypeError):
        return str(val).strip()


def _is_precision_lost(val) -> bool:
    """Return True if val looks like a float-truncated ID (scientific notation or trailing zeros)."""
    if val is None or val == "":
        return False
    s = str(val).strip()
    # Scientific notation from sheet → precision lost
    if 'e' in s.lower():
        return True
    # Normalized value ending in 4+ zeros → likely truncated (e.g. 281750000000000)
    try:
        normalized = str(int(float(s)))
        if len(normalized) >= 10 and normalized.endswith("0000"):
            return True
    except (ValueError, TypeError):
        pass
    return False


def _is_date_on_or_after(date_str: str, cutoff: str) -> bool:
    """Check if date_str (various formats) is on or after cutoff (YYYY-MM-DD)."""
    from datetime import datetime
    s = date_str.strip()
    if not s:
        return True  # allow empty dates through
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%b %d, %Y, %H:%M", "%b %d, %Y",
                "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d") >= cutoff
        except ValueError:
            continue
    return True  # if unparseable, allow through


def _fuzzy_partner_lookup(name: str, emails: dict) -> dict:
    """Look up partner info by name with fuzzy matching fallback."""
    if not name:
        return {}
    key = name.lower().strip()
    # Exact match first
    if key in emails:
        return emails[key]
    # Fuzzy match using difflib
    import difflib
    matches = difflib.get_close_matches(key, emails.keys(), n=1, cutoff=0.95)
    if matches:
        return emails[matches[0]]
    return {}


def _build_partner_id_index(emails: dict, *extra_email_dicts) -> dict:
    """Build reverse lookup {partner_id: {email, partner_id, name}} from email directories.
    Indexes all partner_ids from merged + any extra dicts (e.g. status-only emails)
    so partners with multiple IDs across sources are all reachable."""
    by_id = {}
    for source in [emails, *extra_email_dicts]:
        for name_lower, info in source.items():
            pid = (info.get("partner_id") or "").strip()
            if pid and pid not in by_id:
                by_id[pid] = {**info, "name": info.get("name", name_lower)}
    return by_id


def _lookup_by_partner_id(partner_id: str, by_id: dict) -> dict:
    """Look up partner info by partner_id. Returns {email, partner_id, name} or {}."""
    if not partner_id:
        return {}
    pid = _normalize_partner_id(partner_id)
    return by_id.get(pid, {})

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}},
     allow_headers=["Content-Type", "Authorization"],
     supports_credentials=True)
db.init_db()
db._migrate()

# ── Password protection ──────────────────────────────────────────────────────
APP_USER = os.environ.get("APP_USER", "wiom")
APP_PASS = os.environ.get("APP_PASS", "wiom2026")


def check_auth(username, password):
    return username == APP_USER and password == APP_PASS


def auth_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return Response(
                "Login required.", 401,
                {"WWW-Authenticate": 'Basic realm="WIOM Breach Tracker"'},
            )
        return f(*args, **kwargs)
    return decorated


@app.before_request
def require_login():
    if request.method == 'OPTIONS':
        return  # Let Flask-CORS handle preflight
    auth = request.authorization
    if not auth or not check_auth(auth.username, auth.password):
        resp = Response("Login required.", 401,
                        {"WWW-Authenticate": 'Basic realm="WIOM Breach Tracker"'})
        # Add CORS headers so cross-origin callers can read the error
        origin = request.headers.get('Origin', '*')
        resp.headers['Access-Control-Allow-Origin'] = origin
        resp.headers['Access-Control-Allow-Credentials'] = 'true'
        return resp

# ── SSE broadcast hub ─────────────────────────────────────────────────────────
_sse_clients: list[queue.Queue] = []
_sse_lock = threading.Lock()


def _sse_broadcast(data: dict):
    msg = "data: " + json.dumps(data) + "\n\n"
    with _sse_lock:
        dead = []
        for q in _sse_clients:
            try:
                q.put_nowait(msg)
            except queue.Full:
                dead.append(q)
        for q in dead:
            _sse_clients.remove(q)


@app.route("/api/stream")
def sse_stream():
    def generate():
        q = queue.Queue(maxsize=100)
        with _sse_lock:
            _sse_clients.append(q)
        try:
            yield "data: " + json.dumps({"type": "connected"}) + "\n\n"
            while True:
                try:
                    msg = q.get(timeout=25)
                    yield msg
                except queue.Empty:
                    yield "data: " + json.dumps({"type": "ping"}) + "\n\n"
        except GeneratorExit:
            pass
        finally:
            with _sse_lock:
                try:
                    _sse_clients.remove(q)
                except ValueError:
                    pass

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"},
    )


# ── Background scheduler ──────────────────────────────────────────────────────
def _run_sync(cfg: dict) -> dict:
    """Core sync logic shared by scheduler and manual /api/sync.
    Returns {new_cases, updated, total_fetched, errors, slack_sent}."""
    from metabase import MetabaseClient
    from kapture import KaptureClient

    client = MetabaseClient(
        cfg["metabase_url"], cfg.get("metabase_database_id", ""),
        api_key=cfg.get("metabase_api_key", ""),
        username=cfg.get("metabase_username", ""),
        password=cfg.get("metabase_password", ""),
    )
    kap  = KaptureClient(cfg.get("kapture_auth_header", ""))
    rows = client.run_breach_query()

    new_cases_list = []
    updated_count  = 0
    errors         = []

    for row in rows:
        tid = str(row.get("ticket_id", ""))
        if not tid:
            continue
        existing = sheets_db.get_case(tid)
        sheets_db.upsert_case(row)

        if existing is None:
            # Enrich new case with Kapture right away
            if row.get("kapture_ticket_id"):
                try:
                    raw = kap.fetch_ticket(str(row["kapture_ticket_id"]))
                    if raw:
                        fields = kap.extract_breach_fields(raw)
                        sheets_db.update_kapture_fields(
                            tid, fields["extra_amount"], fields["technician_name"],
                            fields["voluntary_tip"], json.dumps(raw),
                        )
                except Exception as e:
                    errors.append({"ticket_id": tid, "error": str(e)})
            new_cases_list.append(sheets_db.get_case(tid))
        else:
            # Re-fetch Kapture for cases still missing amount
            case_now = sheets_db.get_case(tid)
            if case_now and not case_now.get("extra_amount") and row.get("kapture_ticket_id"):
                try:
                    raw = kap.fetch_ticket(str(row["kapture_ticket_id"]))
                    if raw:
                        fields = kap.extract_breach_fields(raw)
                        sheets_db.update_kapture_fields(
                            tid, fields["extra_amount"], fields["technician_name"],
                            fields["voluntary_tip"], json.dumps(raw),
                        )
                except Exception as e:
                    errors.append({"ticket_id": tid, "error": str(e)})
            updated_count += 1

    if new_cases_list:
        # Push SSE event to all open browsers
        _sse_broadcast({
            "type":       "new_cases",
            "count":      len(new_cases_list),
            "ticket_ids": [c["ticket_id"] for c in new_cases_list if c],
        })

    return {
        "new_cases":     len(new_cases_list),
        "updated":       updated_count,
        "total_fetched": len(rows),
        "errors":        errors,
    }


try:
    from apscheduler.schedulers.background import BackgroundScheduler

    def _sync_b1b4():
        """Sync B1/B4 from Google Sheets into SQLite."""
        try:
            from google_sheets import (fetch_disintermediation_cases, get_all_partner_emails,
                                       fetch_fp4_cases, fetch_churn_feb_cases,
                                       fetch_rohit_call_tagging_cases, fetch_cancelled_calling_cases,
                                       fetch_customer_complaint_cases,
                                       fetch_cx_churn_px_interaction_cases,
                                       fetch_cx_churn_without_tickets_cases)
        except Exception as e:
            app.logger.error(f"B1/B4 import error: {e}")
            return
        # Fetch partner emails once for both B1 and B4
        try:
            emails = get_all_partner_emails()
        except Exception as e:
            app.logger.error(f"Partner emails fetch error: {e}")
            emails = {}
        emails_by_id = _build_partner_id_index(emails)
        # B1 — Source 1: 6mo Churn Sheet1 (existing)
        try:
            cases = fetch_disintermediation_cases()
            cases = [c for c in cases if (c.get("Disintermediation") or "").strip().lower() == "yes"]
            for c in cases:
                partner = (c.get("PARTNER_NAME") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                data = {
                    "lng_nas_id": _normalize_partner_id(c.get("LNG_NAS_ID", "")),
                    "customer_mobile": str(c.get("MOBILE", "") or ""),
                    "expiry_dt": _parse_expiry_date(c.get("EXPIRY_DT", "")),
                    "city": c.get("CITY", ""), "mis_city": c.get("MIS_CITY", ""),
                    "zone": c.get("ZONE", ""), "partner_name": partner,
                    "tenure": c.get("TENURE", ""),
                    "r_oct": c.get("R total (Oct)", ""), "r_nov": c.get("R total (Nov)", ""),
                    "r_dec": c.get("R total (Dec)", ""), "r_jan": c.get("R total (Jan)", ""),
                    "risk_score": c.get("Risk score on wallet activity (Dec+Jan) (Scale 0-3)", ""),
                    "partner_status": c.get("Status", ""), "connected": c.get("Connected", ""),
                    "calling_remarks": c.get("Calling Remarks", ""),
                    "disintermediation": c.get("Disintermediation", ""),
                    "call_recording": c.get("Call Recording", ""), "called_by": c.get("Called By", ""),
                    "call_timestamp": c.get("Call Timestamp (Date)", ""),
                    "calling_status": c.get("Calling Status", ""),
                    "partner_email": email_info.get("email", ""),
                    "partner_id": _normalize_partner_id(email_info.get("partner_id", "")),
                    "source": "churn_logic",
                }
                b1db.upsert_breach1_case(data)
            app.logger.info(f"B1 sync (churn_logic): {len(cases)} cases")
        except Exception as e:
            app.logger.error(f"B1 sync churn_logic error: {e}")
        # B1 — Source 2: 6mo Churn Feb tab
        try:
            cases = fetch_churn_feb_cases()
            cases = [c for c in cases if (c.get("Disintermediation") or "").strip().lower() == "yes"]
            for c in cases:
                partner = (c.get("partner_name") or c.get("PARTNER_NAME") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                raw_pid = c.get("partner_id", "")
                email_pid = _normalize_partner_id(email_info.get("partner_id", ""))
                sheet_pid = _normalize_partner_id(raw_pid)
                # If sheet ID looks precision-lost (scientific notation → trailing zeros), prefer email directory
                partner_id = email_pid if _is_precision_lost(raw_pid) else (sheet_pid or email_pid)
                data = {
                    "lng_nas_id": _normalize_partner_id(c.get("LNG_NAS_ID", "")),
                    "customer_mobile": str(c.get("MOBILE", "") or ""),
                    "expiry_dt": _parse_expiry_date(c.get("EXPIRY_DT", "")),
                    "city": c.get("CITY", ""), "mis_city": c.get("MIS_CITY", ""),
                    "zone": c.get("ZONE", ""), "partner_name": partner,
                    "partner_id": partner_id,
                    "risk_score": c.get("Risk score on wallet activity (Jan+Feb) (Scale 0-3)",
                                        c.get("Risk score on wallet activity (Dec+Jan) (Scale 0-3)", "")),
                    "partner_status": c.get("Status", ""), "connected": c.get("Connected", ""),
                    "calling_remarks": c.get("Calling Remarks", ""),
                    "disintermediation": c.get("Disintermediation", ""),
                    "call_recording": c.get("Call Recording", ""), "called_by": c.get("Called By", ""),
                    "call_timestamp": c.get("Call Date", c.get("Call Timestamp (Date)", "")),
                    "calling_status": c.get("Calling Status", ""),
                    "partner_email": email_info.get("email", ""),
                    "source": "churn_feb",
                }
                b1db.upsert_breach1_case(data)
            app.logger.info(f"B1 sync (churn_feb): {len(cases)} cases")
        except Exception as e:
            app.logger.error(f"B1 sync churn_feb error: {e}")
        # B1 — Source 3: Rohit Call Tagging
        try:
            cases = fetch_rohit_call_tagging_cases()
            count = 0
            for c in cases:
                mobile = c.get("TO_NUMBE", "") or c.get("MOBILE", "")
                partner_id = _normalize_partner_id(c.get("ACCOUNT_ID", ""))
                # Skip rows where both customer_mobile and partner_id are empty
                if not mobile and not partner_id:
                    continue
                # Look up partner name and email from account_id via email directory
                id_info = _lookup_by_partner_id(partner_id, emails_by_id)
                partner = id_info.get("name", "").strip() or (c.get("PARTNER_NAME") or "").strip()
                partner_email = id_info.get("email", "")
                ops_tag = (c.get("Ops Tagging (P1)") or "").strip()
                data = {
                    "lng_nas_id": partner_id,  # use partner_id as dedup key
                    "customer_mobile": mobile,
                    "partner_id": partner_id,
                    "call_recording": c.get("RECORDING_URL", ""),
                    "calling_status": ops_tag,
                    "calling_remarks": ops_tag,
                    "call_timestamp": c.get("CREATED_AT", ""),
                    "partner_name": partner,
                    "partner_email": partner_email,
                    "source": "rohit_call_tagging",
                }
                b1db.upsert_breach1_case(data)
                count += 1
            app.logger.info(f"B1 sync (rohit_call_tagging): {count} cases")
        except Exception as e:
            app.logger.error(f"B1 sync rohit_call_tagging error: {e}")
        # B1 — Source 4: Cancelled Calling
        try:
            cases = fetch_cancelled_calling_cases()
            # Filter already done in fetch_cancelled_calling_cases (Bucketing == Disintermediation)
            for c in cases:
                partner = (c.get("Partner Name(if Disintermediation)") or c.get("Partner Name (if Disintermediation)") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                pid = c.get("Latest Partner ID", "") or _normalize_partner_id(email_info.get("partner_id", ""))
                data = {
                    "customer_mobile": c.get("Mobile", ""),
                    "lng_nas_id": "",
                    "partner_id": pid,
                    "partner_name": partner,
                    "call_recording": c.get("Call Recording", ""),
                    "calling_remarks": c.get("Remarks", ""),
                    "calling_status": c.get("Bucketing", ""),
                    "report_text": c.get("App Reason", ""),
                    "connected": c.get("Call connected", ""),
                    "call_timestamp": c.get("Call Date", ""),
                    "cancelled_time": c.get("Cancelled Time", ""),
                    "partner_email": email_info.get("email", ""),
                    "source": "cancelled_calling",
                }
                b1db.upsert_breach1_case(data)
            app.logger.info(f"B1 sync (cancelled_calling): {len(cases)} cases")
        except Exception as e:
            app.logger.error(f"B1 sync cancelled_calling error: {e}")
        # B1 — Source 5: Customer Complaints
        try:
            cases = fetch_customer_complaint_cases()
            count = 0
            for c in cases:
                kapture_no = (c.get("Kapture Ticket No.") or "").strip()
                if kapture_no == "1" or (kapture_no and not kapture_no.replace(".", "").isdigit()):
                    continue
                # Filter: only cases from March 2026 onwards
                created_date = (c.get("Created Date") or "").strip()
                if created_date and not _is_date_on_or_after(created_date, "2026-03-01"):
                    continue
                partner = (c.get("Partner name") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                data = {
                    "customer_mobile": c.get("Customer Phone", ""),
                    "partner_name": partner,
                    "partner_id": _normalize_partner_id(email_info.get("partner_id", "")),
                    "calling_status": (c.get("Leakage Category") or "").strip(),
                    "calling_remarks": c.get("Leakage confirmation", ""),
                    "call_recording": c.get("Call Recording proof", ""),
                    "report_text": c.get("Ticket URL (Comment)", ""),
                    "lng_nas_id": kapture_no,
                    "call_timestamp": created_date,
                    "partner_email": email_info.get("email", ""),
                    "source": "customer_complaint",
                }
                b1db.upsert_breach1_case(data)
                count += 1
            app.logger.info(f"B1 sync (customer_complaint): {count} cases")
        except Exception as e:
            app.logger.error(f"B1 sync customer_complaint error: {e}")
        # B1 — Source 6: Cx Churn After Px Interaction
        try:
            cases = fetch_cx_churn_px_interaction_cases()
            for c in cases:
                partner = str(c.get("partner_name") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                pid = _normalize_partner_id(c.get("partner_id", "")) or _normalize_partner_id(email_info.get("partner_id", ""))
                data = {
                    "customer_mobile": _normalize_partner_id(c.get("customer_mobile", "")),
                    "lng_nas_id": str(c.get("device_id") or "").strip(),
                    "partner_id": pid,
                    "partner_name": partner,
                    "connected": str(c.get("Called") or "").strip(),
                    "calling_remarks": str(c.get("action_notes") or "").strip(),
                    "disintermediation": str(c.get("Disintermediation_status") or "").strip(),
                    "call_recording": str(c.get("Call Recording") or "").strip(),
                    "call_timestamp": _parse_expiry_date(c.get("Call_date", "")),
                    "city": str(c.get("cluster") or "").strip(),
                    "expiry_dt": _parse_expiry_date(c.get("plan_end_time", "")),
                    "partner_email": email_info.get("email", ""),
                    "source": "cx_churn_px_interaction",
                }
                b1db.upsert_breach1_case(data)
            app.logger.info(f"B1 sync (cx_churn_px_interaction): {len(cases)} cases")
        except Exception as e:
            app.logger.error(f"B1 sync cx_churn_px_interaction error: {e}")
        # B1 — Source 7: Cx Churn Without Service Tickets
        try:
            cases = fetch_cx_churn_without_tickets_cases()
            for c in cases:
                partner = str(c.get("partner_name") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                pid = _normalize_partner_id(c.get("partner_id", "")) or _normalize_partner_id(email_info.get("partner_id", ""))
                data = {
                    "customer_mobile": str(c.get("MOBILE") or "").strip(),
                    "lng_nas_id": str(c.get("NASID") or "").strip(),
                    "partner_id": pid,
                    "partner_name": partner,
                    "connected": str(c.get("Call status") or "").strip(),
                    "calling_remarks": str(c.get("action_notes") or "").strip(),
                    "disintermediation": str(c.get("Disintermediation") or "").strip(),
                    "call_recording": str(c.get("Call Recording") or "").strip(),
                    "call_timestamp": str(c.get("Call_date") or "").strip(),
                    "expiry_dt": _parse_expiry_date(c.get("PLAN_END", "")),
                    "partner_email": email_info.get("email", ""),
                    "source": "cx_churn_no_tickets",
                }
                b1db.upsert_breach1_case(data)
            app.logger.info(f"B1 sync (cx_churn_no_tickets): {len(cases)} cases")
        except Exception as e:
            app.logger.error(f"B1 sync cx_churn_no_tickets error: {e}")
        # B4
        try:
            cases = fetch_fp4_cases()
            for c in cases:
                partner_name = (c.get("Partner Name") or "").strip()
                partner_id = (c.get("Partner Id") or "").strip()
                email_info = emails.get(partner_name.lower(), {})
                data = {
                    "partner_id": partner_id, "partner_name": partner_name,
                    "customer_details": (c.get("Customer Details") or "").strip(),
                    "principle_broken": (c.get("Principle Broken") or "").strip(),
                    "device_id": (c.get("Device Id") or "").strip(),
                    "date_reported": (c.get("Date Reported") or "").strip(),
                    "reporting_channel": (c.get("Reporting Channel") or "").strip(),
                    "penalty_amount": (c.get("Penalty Amount") or "").strip(),
                    "penalty_done": (c.get("Penalty Done") or "").strip(),
                    "penalty_done_date": (c.get("Penalty Done Date") or "").strip(),
                    "partner_email_comms": (c.get("Partner Email Comms Done") or "").strip(),
                    "email_date": (c.get("Email Date") or "").strip(),
                    "whatsapp_comms": (c.get("Partner Text/Whatsapp Comms Done") or "").strip(),
                    "whatsapp_date": (c.get("Whatsapp Date") or "").strip(),
                    "partner_mobile": (c.get("Partner Mobile") or "").strip(),
                    "link": (c.get("Link") or "").strip(),
                    "comments": (c.get("Comments") or "").strip(),
                    "partner_email": email_info.get("email", ""),
                }
                if not data["partner_id"] and not data["partner_name"] and not data["device_id"]:
                    continue
                db.upsert_breach4_case(data)
            app.logger.info(f"B4 sync: {len(cases)} cases")
        except Exception as e:
            app.logger.error(f"B4 sync error: {e}")

    def _auto_sync():
        cfg = config.load()
        if cfg.get("metabase_url") and (cfg.get("metabase_api_key") or cfg.get("metabase_username")) and cfg.get("metabase_database_id"):
            try:
                _run_sync(cfg)
            except Exception as e:
                app.logger.error(f"Auto-sync B2 error: {e}")
        _sync_b1b4()

    def _slack_new_b2_alert():
        try:
            from datetime import datetime, timedelta
            webhook = config.load().get("slack_webhook_url", "")
            if not webhook:
                return
            cases = sheets_db.get_all_cases(state="detected")
            # Only cases from today or last 2 days, with amount
            cutoff = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d")
            recent = []
            for c in cases:
                if not c.get("extra_amount"):
                    continue
                ts = c.get("ticket_added_time_ist") or ""
                date_part = ts.replace("T", " ")[:10]
                if date_part >= cutoff:
                    recent.append(c)
            if not recent:
                return
            lines = [f"<!channel>\n:rotating_light: *{len(recent)} Breach 2 Case(s) — Pending Action*\n"]
            for c in recent:
                amt = f"\u20b9{int(c['extra_amount'])}"
                ts_raw = c.get("ticket_added_time_ist") or ""
                created = ts_raw.replace("T", " ")[:16] if ts_raw else "—"
                lines.append(
                    f"\u2022 `{c.get('kapture_ticket_id') or c['ticket_id']}` | "
                    f"{c.get('current_partner_name', '\u2014')} | {c.get('zone', '\u2014')} | "
                    f"*{amt}* | Created: {created}"
                )
            lines.append("\n<https://shivanksood-prog.github.io/breach_tracker/|Open Dashboard>")
            actions.send_to_slack(webhook, "\n".join(lines))
        except Exception as e:
            app.logger.error(f"Slack new B2 alert error: {e}")

    from datetime import datetime, timedelta
    scheduler = BackgroundScheduler(daemon=True)
    # Defer first _auto_sync by 2 min so it doesn't overlap with the startup job
    scheduler.add_job(_auto_sync, "interval", minutes=2, id="auto_sync",
                      next_run_time=datetime.now() + timedelta(minutes=2))
    scheduler.add_job(_slack_new_b2_alert, "interval", minutes=15, id="b2_new_alert")
    # B1/B4 sync: every 1 hour (multiple sources are slow), plus immediate startup
    scheduler.add_job(_sync_b1b4, "interval", minutes=60, id="b1b4_sync",
                      next_run_time=datetime.now() + timedelta(minutes=60))
    scheduler.add_job(_sync_b1b4, id="b1b4_startup")
    scheduler.start()
except ImportError:
    pass

# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ── Cases ─────────────────────────────────────────────────────────────────────

@app.route("/api/cases")
def get_cases():
    state  = request.args.get("state", "all")
    zone   = request.args.get("zone", "all")
    search = request.args.get("search", "").strip() or None
    cases  = sheets_db.get_all_cases(state=state, zone=zone, search=search)
    return jsonify(cases)


@app.route("/api/cases/summary")
def get_summary():
    return jsonify(sheets_db.get_summary())


@app.route("/api/cases/zones")
def get_zones():
    return jsonify(sheets_db.get_all_zones())


@app.route("/api/cases/<ticket_id>")
def get_case(ticket_id):
    case = sheets_db.get_case(ticket_id)
    if not case:
        return jsonify({"error": "Not found"}), 404
    return jsonify(case)


# ── Sync ──────────────────────────────────────────────────────────────────────

@app.route("/api/sync", methods=["POST"])
def sync():
    cfg = config.load()
    if not cfg.get("metabase_url"):
        return jsonify({"error": "Metabase not configured. Go to Settings."}), 400
    if not cfg.get("metabase_database_id"):
        return jsonify({"error": "Database ID not set. Go to Settings → Discover Databases."}), 400
    try:
        result = _run_sync(cfg)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cases/<ticket_id>/fetch-kapture", methods=["POST"])
def fetch_kapture_single(ticket_id):
    cfg  = config.load()
    case = sheets_db.get_case(ticket_id)
    if not case:
        return jsonify({"error": "Case not found"}), 404
    kapture_id = case.get("kapture_ticket_id")
    if not kapture_id:
        return jsonify({"error": "No Kapture ticket ID on this case"}), 400
    try:
        from kapture import KaptureClient
        kap = KaptureClient(cfg.get("kapture_auth_header", ""))
        raw = kap.fetch_ticket(str(kapture_id))
        if raw:
            fields = kap.extract_breach_fields(raw)
            sheets_db.update_kapture_fields(
                ticket_id, fields["extra_amount"], fields["technician_name"],
                fields["voluntary_tip"], json.dumps(raw),
            )
        return jsonify(sheets_db.get_case(ticket_id))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Actions ───────────────────────────────────────────────────────────────────



# ── Customer Comms ────────────────────────────────────────────────────────────

@app.route("/api/customer-comms")
def get_customer_comms():
    return jsonify(sheets_db.get_pending_comms())


@app.route("/api/customer-comms/<ticket_id>/log", methods=["POST"])
def log_comms(ticket_id):
    body = request.json or {}
    notes = body.get("notes", "")
    ok = sheets_db.advance_state(ticket_id, "customer_comms", comms_notes=notes)
    if ok:
        return jsonify({"ok": True, "case": sheets_db.get_case(ticket_id)})
    return jsonify({"error": "Cannot advance state"}), 400


@app.route("/api/cases/<ticket_id>/advance", methods=["POST"])
def advance_case(ticket_id):
    body      = request.json or {}
    new_state = body.get("state")
    if not new_state:
        return jsonify({"error": "Missing state"}), 400
    ok = sheets_db.advance_state(ticket_id, new_state)
    if ok:
        return jsonify({"ok": True, "case": sheets_db.get_case(ticket_id)})
    return jsonify({"error": "Invalid state transition"}), 400


@app.route("/api/cases/<ticket_id>/undo", methods=["POST"])
def undo_case(ticket_id):
    ok = sheets_db.undo_state(ticket_id)
    if ok:
        return jsonify({"ok": True, "case": sheets_db.get_case(ticket_id)})
    return jsonify({"error": "Nothing to undo"}), 400




# ── CSV Downloads ─────────────────────────────────────────────────────────────

@app.route("/api/download/refund-csv")
def dl_refund():
    cases   = sheets_db.get_all_cases(state="detected")
    csv_str = actions.generate_refund_csv(cases)
    return Response(csv_str, mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=refund_cases.csv"})


@app.route("/api/csv/refund", methods=["POST"])
def csv_refund():
    tids  = (request.json or {}).get("ticket_ids", [])
    cases = ([sheets_db.get_case(t) for t in tids] if tids
             else sheets_db.get_all_cases(state="detected"))
    cases = [c for c in cases if c]
    return jsonify({"csv": actions.generate_refund_csv(cases), "count": len(cases)})


@app.route("/api/csv/penalty", methods=["POST"])
def csv_penalty():
    tids  = (request.json or {}).get("ticket_ids", [])
    cases = ([sheets_db.get_case(t) for t in tids] if tids
             else sheets_db.get_all_cases(state="customer_comms"))
    cases = [c for c in cases if c]
    xlsx_bytes = actions.generate_penalty_xlsx(cases)
    return Response(xlsx_bytes,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=partner_penalty.xlsx"})


@app.route("/api/csv/partner-comms", methods=["POST"])
def csv_partner_comms_bulk():
    tids  = (request.json or {}).get("ticket_ids", [])
    cases = ([sheets_db.get_case(t) for t in tids] if tids
             else sheets_db.get_all_cases(state="partner_penalty"))
    cases = [c for c in cases if c]
    return jsonify({"csv": actions.generate_partner_comms_csv(cases), "count": len(cases)})


@app.route("/api/download/penalty-csv")
def dl_penalty():
    cases = sheets_db.get_all_cases(state="customer_comms")
    xlsx_bytes = actions.generate_penalty_xlsx(cases)
    return Response(xlsx_bytes,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=partner_penalty.xlsx"})


@app.route("/api/download/partner-comms-csv")
def dl_partner_comms():
    cases   = sheets_db.get_all_cases(state="partner_penalty")
    csv_str = actions.generate_partner_comms_csv(cases)
    return Response(csv_str, mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=partner_comms.csv"})


def _calc_tat_minutes(t1_str, t2_str):
    """Calculate TAT in minutes between two timestamp strings."""
    try:
        from datetime import datetime as _dt
        t1 = _dt.strptime(t1_str.replace("T", " ")[:19], "%Y-%m-%d %H:%M:%S")
        t2 = _dt.strptime(t2_str.replace("T", " ")[:19], "%Y-%m-%d %H:%M:%S")
        return round((t2 - t1).total_seconds() / 60, 1)
    except Exception:
        return None


# ── Refund Upload ────────────────────────────────────────────────────────────

@app.route("/api/upload/refund-status", methods=["POST"])
def upload_refund_status():
    import csv
    import io
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        text = f.read().decode("utf-8-sig")
    except Exception:
        return jsonify({"error": "Could not read file as UTF-8"}), 400
    reader = csv.DictReader(io.StringIO(text))
    matched = []
    unmatched = []
    for row in reader:
        phone = (row.get("Contact Phone Number") or "").strip()
        payout_id = (row.get("Payout Link ID") or "").strip()
        if not phone or not payout_id:
            continue
        # Normalize: strip leading +91 or 91 prefix to get 10-digit number
        clean = phone.lstrip("+")
        if clean.startswith("91") and len(clean) > 10:
            clean = clean[2:]
        result = sheets_db.mark_refunded_by_mobile(clean, payout_id)
        if result["matched"]:
            matched.append(result)
        else:
            # Also try with original phone value
            if clean != phone:
                result2 = sheets_db.mark_refunded_by_mobile(phone, payout_id)
                if result2["matched"]:
                    matched.append(result2)
                    continue
            unmatched.append({"mobile": phone, "payout_link_id": payout_id})
    # ── Slack notification for matched refund cases ──
    if matched:
        try:
            slack_lines = ["<!channel>\n*Refund Processed — Customer Comms Pending*\n"]
            for m in matched:
                tid = m.get("ticket_id", "")
                case_data = sheets_db.get_case(tid)
                if not case_data:
                    continue
                kapture_id = case_data.get("kapture_ticket_id") or "N/A"
                partner_name = case_data.get("current_partner_name") or "N/A"
                mobile = case_data.get("customer_mobile") or "N/A"
                extra_amt = case_data.get("extra_amount")
                amt_str = f"₹{extra_amt}" if extra_amt is not None else "N/A"
                t1 = case_data.get("ticket_added_time_ist") or ""
                t2 = case_data.get("customer_refunded_at") or ""
                tat = _calc_tat_minutes(t1, t2)
                tat_str = f"{tat} min" if tat is not None else "N/A"
                slack_lines.append(
                    f"• Kapture {kapture_id} | {partner_name} | {mobile} | {amt_str} | TAT {tat_str}"
                )
            slack_lines.append("\n<https://shivanksood-prog.github.io/breach_tracker/|Open Customer Comms Tab>")
            webhook_url = config.get("slack_webhook_url", "")
            if webhook_url and len(slack_lines) > 1:
                actions.send_to_slack(webhook_url, "\n".join(slack_lines))
        except Exception:
            pass  # Don't break upload response on Slack failure
    return jsonify({
        "ok": True,
        "matched_count": len(matched),
        "unmatched_count": len(unmatched),
        "matched": matched,
        "unmatched": unmatched,
    })


@app.route("/api/upload/penalty-status", methods=["POST"])
def upload_penalty_status():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400

    rows = []
    fname = (f.filename or "").lower()
    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        # Excel upload
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(f.read()), read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        wb.close()
    else:
        # CSV fallback
        import csv, io
        try:
            text = f.read().decode("utf-8-sig")
        except Exception:
            return jsonify({"error": "Could not read file"}), 400
        rows = list(csv.DictReader(io.StringIO(text)))

    matched = []
    unmatched = []
    for row in rows:
        partner_id = str(row.get("AccountId") or row.get("Partner Id") or "").strip()
        if not partner_id:
            continue
        # Skip rows where Process Status is not "Yes"
        status = str(row.get("Process Status") or "").strip().lower()
        if status and status != "yes":
            unmatched.append({"partner_id": partner_id, "reason": row.get("Reason", "")})
            continue
        result = sheets_db.mark_penalty_by_upload(partner_id)
        if result["matched"]:
            matched.append(result)
        else:
            unmatched.append({"partner_id": partner_id})
    return jsonify({
        "ok": True,
        "matched_count": len(matched),
        "unmatched_count": len(unmatched),
        "matched": matched,
        "unmatched": unmatched,
    })


@app.route("/api/cases/<ticket_id>/confirm-comms", methods=["POST"])
def confirm_comms(ticket_id):
    body = request.json or {}
    notes = body.get("notes", "")
    ok = sheets_db.advance_state(ticket_id, "customer_comms", comms_notes=notes)
    if ok:
        return jsonify({"ok": True, "case": sheets_db.get_case(ticket_id)})
    return jsonify({"error": "Cannot confirm comms for this case"}), 400


# ── Visibility Dashboard ──────────────────────────────────────────────────────

@app.route("/api/cases/visibility")
def visibility_matrix():
    try:
        return jsonify(sheets_db.get_visibility_matrix())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Fraud Tracker ─────────────────────────────────────────────────────────────

@app.route("/api/cases/repeat-customers")
def repeat_customers():
    return jsonify(sheets_db.get_repeat_customers())


# ── FP2 Partner Email ─────────────────────────────────────────────────────────

@app.route("/api/breach2/template")
def breach2_template():
    from email_sender import get_fp2_template_info
    return jsonify(get_fp2_template_info())


@app.route("/api/breach2/preview-email", methods=["POST"])
def breach2_preview():
    from email_sender import render_fp2_email
    body = request.json or {}
    language = body.get("language", "both")
    selected_vars = body.get("selected_vars", [])
    values = body.get("values", {})
    try:
        result = render_fp2_email(language, selected_vars, values)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/breach2/send-email", methods=["POST"])
def breach2_send_email():
    from email_sender import render_fp2_email, send_email
    from google_sheets import get_all_partner_emails
    body = request.json or {}
    ticket_ids = body.get("ticket_ids", [])
    language = body.get("language", "both")
    selected_vars = body.get("selected_vars", [])
    values = body.get("values", {})
    test_email = body.get("test_email", "").strip()
    is_test = bool(test_email)

    partner_emails = get_all_partner_emails()
    results = []
    for tid in ticket_ids:
        case = sheets_db.get_case(tid)
        if not case:
            results.append({"ticket_id": tid, "ok": False, "error": "Case not found"})
            continue

        vars_filled = dict(values)
        # Auto-fill FP2 variables from case data
        if not vars_filled.get("MASKED_NUMBER"):
            mobile = case.get("customer_mobile", "")
            vars_filled["MASKED_NUMBER"] = (mobile[:4] + "XXXXXX") if len(mobile) >= 4 else mobile
        if not vars_filled.get("AMOUNT_COLLECTED"):
            vars_filled["AMOUNT_COLLECTED"] = str(int(float(case.get("extra_amount") or 0)))
        if not vars_filled.get("AMOUNT_REFUNDED"):
            vars_filled["AMOUNT_REFUNDED"] = str(int(float(case.get("extra_amount") or 0)))
        if not vars_filled.get("AMOUNT_RECOVERED"):
            vars_filled["AMOUNT_RECOVERED"] = str(int(float(case.get("extra_amount") or 0)))

        rendered = render_fp2_email(language, selected_vars, vars_filled)
        partner_name = (case.get("current_partner_name") or "").strip()
        email_info = partner_emails.get(partner_name.lower(), {})
        recipient = test_email if is_test else email_info.get("email", "")

        if not recipient:
            results.append({"ticket_id": tid, "ok": False, "error": "No partner email"})
            continue

        send_result = send_email(recipient, rendered["subject"], rendered["body_text"], rendered["body_html"])
        results.append({"ticket_id": tid, "partner": partner_name, "recipient": recipient, **send_result})

    return jsonify({"results": results, "total": len(results),
                    "sent": sum(1 for r in results if r.get("ok"))})


# ── Breach 1 (Disintermediation) ──────────────────────────────────────────────

@app.route("/api/breach1/sync", methods=["POST"])
def breach1_sync():
    """Import disintermediation cases from all sources and match partner emails."""
    try:
        from google_sheets import (fetch_disintermediation_cases, get_all_partner_emails,
                                   fetch_churn_feb_cases, fetch_rohit_call_tagging_cases,
                                   fetch_cancelled_calling_cases, fetch_customer_complaint_cases)
        emails = get_all_partner_emails()
        emails_by_id = _build_partner_id_index(emails)
        total_new = 0
        total_updated = 0
        total_fetched = 0
        source_counts = {}

        def _count_and_upsert(data):
            nonlocal total_new, total_updated
            existing = b1db.get_breach1_cases(search=data.get("lng_nas_id") or data.get("customer_mobile", ""))
            existing = [e for e in existing if e.get("customer_mobile") == data.get("customer_mobile")]
            if not existing:
                total_new += 1
            else:
                total_updated += 1
            b1db.upsert_breach1_case(data)

        # Source 1: 6mo Churn Sheet1
        try:
            cases = fetch_disintermediation_cases()
            cases = [c for c in cases if (c.get("Disintermediation") or "").strip().lower() == "yes"]
            for c in cases:
                partner = (c.get("PARTNER_NAME") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                data = {
                    "lng_nas_id": _normalize_partner_id(c.get("LNG_NAS_ID", "")),
                    "customer_mobile": str(c.get("MOBILE", "") or ""),
                    "expiry_dt": _parse_expiry_date(c.get("EXPIRY_DT", "")),
                    "city": c.get("CITY", ""), "mis_city": c.get("MIS_CITY", ""),
                    "zone": c.get("ZONE", ""), "partner_name": partner,
                    "tenure": c.get("TENURE", ""),
                    "r_oct": c.get("R total (Oct)", ""), "r_nov": c.get("R total (Nov)", ""),
                    "r_dec": c.get("R total (Dec)", ""), "r_jan": c.get("R total (Jan)", ""),
                    "risk_score": c.get("Risk score on wallet activity (Dec+Jan) (Scale 0-3)", ""),
                    "partner_status": c.get("Status", ""), "connected": c.get("Connected", ""),
                    "calling_remarks": c.get("Calling Remarks", ""),
                    "disintermediation": c.get("Disintermediation", ""),
                    "call_recording": c.get("Call Recording", ""), "called_by": c.get("Called By", ""),
                    "call_timestamp": c.get("Call Timestamp (Date)", ""),
                    "calling_status": c.get("Calling Status", ""),
                    "partner_email": email_info.get("email", ""),
                    "partner_id": _normalize_partner_id(email_info.get("partner_id", "")),
                    "source": "churn_logic",
                }
                _count_and_upsert(data)
            source_counts["churn_logic"] = len(cases)
            total_fetched += len(cases)
        except Exception as e:
            source_counts["churn_logic"] = f"error: {e}"

        # Source 2: 6mo Churn Feb
        try:
            cases = fetch_churn_feb_cases()
            cases = [c for c in cases if (c.get("Disintermediation") or "").strip().lower() == "yes"]
            for c in cases:
                partner = (c.get("partner_name") or c.get("PARTNER_NAME") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                raw_pid = c.get("partner_id", "")
                email_pid = _normalize_partner_id(email_info.get("partner_id", ""))
                sheet_pid = _normalize_partner_id(raw_pid)
                # If sheet ID looks precision-lost (scientific notation → trailing zeros), prefer email directory
                partner_id = email_pid if _is_precision_lost(raw_pid) else (sheet_pid or email_pid)
                data = {
                    "lng_nas_id": _normalize_partner_id(c.get("LNG_NAS_ID", "")),
                    "customer_mobile": str(c.get("MOBILE", "") or ""),
                    "expiry_dt": _parse_expiry_date(c.get("EXPIRY_DT", "")),
                    "city": c.get("CITY", ""), "mis_city": c.get("MIS_CITY", ""),
                    "zone": c.get("ZONE", ""), "partner_name": partner,
                    "partner_id": partner_id,
                    "risk_score": c.get("Risk score on wallet activity (Jan+Feb) (Scale 0-3)",
                                        c.get("Risk score on wallet activity (Dec+Jan) (Scale 0-3)", "")),
                    "partner_status": c.get("Status", ""), "connected": c.get("Connected", ""),
                    "calling_remarks": c.get("Calling Remarks", ""),
                    "disintermediation": c.get("Disintermediation", ""),
                    "call_recording": c.get("Call Recording", ""), "called_by": c.get("Called By", ""),
                    "call_timestamp": c.get("Call Date", c.get("Call Timestamp (Date)", "")),
                    "calling_status": c.get("Calling Status", ""),
                    "partner_email": email_info.get("email", ""),
                    "source": "churn_feb",
                }
                _count_and_upsert(data)
            source_counts["churn_feb"] = len(cases)
            total_fetched += len(cases)
        except Exception as e:
            source_counts["churn_feb"] = f"error: {e}"

        # Source 3: Rohit Call Tagging
        try:
            cases = fetch_rohit_call_tagging_cases()
            count = 0
            for c in cases:
                mobile = c.get("TO_NUMBE", "") or c.get("MOBILE", "")
                partner_id = _normalize_partner_id(c.get("ACCOUNT_ID", ""))
                # Skip rows where both customer_mobile and partner_id are empty
                if not mobile and not partner_id:
                    continue
                # Look up partner name and email from account_id via email directory
                id_info = _lookup_by_partner_id(partner_id, emails_by_id)
                partner = id_info.get("name", "").strip() or (c.get("PARTNER_NAME") or "").strip()
                partner_email = id_info.get("email", "")
                ops_tag = (c.get("Ops Tagging (P1)") or "").strip()
                data = {
                    "lng_nas_id": partner_id,  # use partner_id as dedup key
                    "customer_mobile": mobile,
                    "partner_id": partner_id,
                    "call_recording": c.get("RECORDING_URL", ""),
                    "calling_status": ops_tag,
                    "calling_remarks": ops_tag,
                    "call_timestamp": c.get("CREATED_AT", ""),
                    "partner_name": partner,
                    "partner_email": partner_email,
                    "source": "rohit_call_tagging",
                }
                _count_and_upsert(data)
                count += 1
            source_counts["rohit_call_tagging"] = count
            total_fetched += count
        except Exception as e:
            source_counts["rohit_call_tagging"] = f"error: {e}"

        # Source 4: Cancelled Calling
        try:
            cases = fetch_cancelled_calling_cases()
            # Filter already done in fetch_cancelled_calling_cases (Bucketing == Disintermediation)
            for c in cases:
                partner = (c.get("Partner Name(if Disintermediation)") or c.get("Partner Name (if Disintermediation)") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                pid = c.get("Latest Partner ID", "") or _normalize_partner_id(email_info.get("partner_id", ""))
                data = {
                    "customer_mobile": c.get("Mobile", ""),
                    "lng_nas_id": "",
                    "partner_id": pid,
                    "partner_name": partner,
                    "call_recording": c.get("Call Recording", ""),
                    "calling_remarks": c.get("Remarks", ""),
                    "calling_status": c.get("Bucketing", ""),
                    "report_text": c.get("App Reason", ""),
                    "connected": c.get("Call connected", ""),
                    "call_timestamp": c.get("Call Date", ""),
                    "cancelled_time": c.get("Cancelled Time", ""),
                    "partner_email": email_info.get("email", ""),
                    "source": "cancelled_calling",
                }
                _count_and_upsert(data)
            source_counts["cancelled_calling"] = len(cases)
            total_fetched += len(cases)
        except Exception as e:
            source_counts["cancelled_calling"] = f"error: {e}"

        # Source 5: Customer Complaints
        try:
            cases = fetch_customer_complaint_cases()
            # Filter already done in fetch_customer_complaint_cases (Leakage Category == disintermediation)
            count = 0
            for c in cases:
                kapture_no = (c.get("Kapture Ticket No.") or "").strip()
                if kapture_no == "1" or (kapture_no and not kapture_no.replace(".", "").isdigit()):
                    continue
                created_date = (c.get("Created Date") or "").strip()
                if created_date and not _is_date_on_or_after(created_date, "2026-03-01"):
                    continue
                partner = (c.get("Partner name") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                data = {
                    "customer_mobile": c.get("Customer Phone", ""),
                    "partner_name": partner,
                    "partner_id": _normalize_partner_id(email_info.get("partner_id", "")),
                    "calling_status": (c.get("Leakage Category") or "").strip(),
                    "calling_remarks": c.get("Leakage confirmation", ""),
                    "call_recording": c.get("Call Recording proof", ""),
                    "report_text": c.get("Ticket URL (Comment)", ""),
                    "lng_nas_id": kapture_no,
                    "call_timestamp": created_date,
                    "partner_email": email_info.get("email", ""),
                    "source": "customer_complaint",
                }
                _count_and_upsert(data)
                count += 1
            source_counts["customer_complaint"] = count
            total_fetched += count
        except Exception as e:
            source_counts["customer_complaint"] = f"error: {e}"

        # Source 6: Cx Churn After Px Interaction
        try:
            cases = fetch_cx_churn_px_interaction_cases()
            for c in cases:
                partner = str(c.get("partner_name") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                pid = _normalize_partner_id(c.get("partner_id", "")) or _normalize_partner_id(email_info.get("partner_id", ""))
                data = {
                    "customer_mobile": _normalize_partner_id(c.get("customer_mobile", "")),
                    "lng_nas_id": str(c.get("device_id") or "").strip(),
                    "partner_id": pid,
                    "partner_name": partner,
                    "connected": str(c.get("Called") or "").strip(),
                    "calling_remarks": str(c.get("action_notes") or "").strip(),
                    "disintermediation": str(c.get("Disintermediation_status") or "").strip(),
                    "call_recording": str(c.get("Call Recording") or "").strip(),
                    "call_timestamp": _parse_expiry_date(c.get("Call_date", "")),
                    "city": str(c.get("cluster") or "").strip(),
                    "expiry_dt": _parse_expiry_date(c.get("plan_end_time", "")),
                    "partner_email": email_info.get("email", ""),
                    "source": "cx_churn_px_interaction",
                }
                _count_and_upsert(data)
            source_counts["cx_churn_px_interaction"] = len(cases)
            total_fetched += len(cases)
        except Exception as e:
            source_counts["cx_churn_px_interaction"] = f"error: {e}"

        # Source 7: Cx Churn Without Service Tickets
        try:
            cases = fetch_cx_churn_without_tickets_cases()
            for c in cases:
                partner = str(c.get("partner_name") or "").strip()
                email_info = _fuzzy_partner_lookup(partner, emails)
                pid = _normalize_partner_id(c.get("partner_id", "")) or _normalize_partner_id(email_info.get("partner_id", ""))
                data = {
                    "customer_mobile": str(c.get("MOBILE") or "").strip(),
                    "lng_nas_id": str(c.get("NASID") or "").strip(),
                    "partner_id": pid,
                    "partner_name": partner,
                    "connected": str(c.get("Call status") or "").strip(),
                    "calling_remarks": str(c.get("action_notes") or "").strip(),
                    "disintermediation": str(c.get("Disintermediation") or "").strip(),
                    "call_recording": str(c.get("Call Recording") or "").strip(),
                    "call_timestamp": str(c.get("Call_date") or "").strip(),
                    "expiry_dt": _parse_expiry_date(c.get("PLAN_END", "")),
                    "partner_email": email_info.get("email", ""),
                    "source": "cx_churn_no_tickets",
                }
                _count_and_upsert(data)
            source_counts["cx_churn_no_tickets"] = len(cases)
            total_fetched += len(cases)
        except Exception as e:
            source_counts["cx_churn_no_tickets"] = f"error: {e}"

        return jsonify({"ok": True, "new_cases": total_new, "updated": total_updated,
                        "total_fetched": total_fetched, "by_source": source_counts})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/breach1/cases")
def breach1_cases():
    partner = request.args.get("partner", "all")
    zone = request.args.get("zone", "all")
    status = request.args.get("status", "all")
    email_state = request.args.get("email_state", "all")
    source = request.args.get("source", "all")
    action_type = request.args.get("action_type", "all")
    search = request.args.get("search", "").strip() or None
    return jsonify(b1db.get_breach1_cases(partner=partner, zone=zone, status=status,
                                        email_state=email_state, search=search,
                                        source=source, action_type=action_type))


@app.route("/api/breach1/cases/<case_id>")
def breach1_case(case_id):
    case = b1db.get_breach1_case(case_id)
    if not case:
        return jsonify({"error": "Not found"}), 404
    return jsonify(case)


@app.route("/api/breach1/summary")
def breach1_summary():
    return jsonify(b1db.get_breach1_summary())


@app.route("/api/breach1/dashboard")
def breach1_dashboard():
    return jsonify(b1db.get_breach1_dashboard())


@app.route("/api/breach1/partners")
def breach1_partners():
    return jsonify(b1db.get_breach1_partners())


@app.route("/api/breach1/zones")
def breach1_zones():
    return jsonify(b1db.get_breach1_zones())


@app.route("/api/breach1/template")
def breach1_template():
    from email_sender import get_template_info
    return jsonify(get_template_info())


@app.route("/api/breach1/preview-email", methods=["POST"])
def breach1_preview():
    from email_sender import render_email
    body = request.json or {}
    language = body.get("language", "both")
    selected_vars = body.get("selected_vars", [])
    values = body.get("values", {})
    try:
        result = render_email(language, selected_vars, values)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/breach1/send-email", methods=["POST"])
def breach1_send_email():
    from email_sender import render_email, send_email
    body = request.json or {}
    case_ids = body.get("case_ids", [])
    language = body.get("language", "both")
    selected_vars = body.get("selected_vars", [])
    values = body.get("values", {})
    test_email = body.get("test_email", "").strip()
    is_test = bool(test_email)

    results = []
    for cid in case_ids:
        case = b1db.get_breach1_case(cid)
        if not case:
            results.append({"case_id": cid, "ok": False, "error": "Case not found"})
            continue

        # Auto-fill from case data where possible
        vars_filled = dict(values)
        if "LAST_SERVICE_DATE" in selected_vars and not vars_filled.get("LAST_SERVICE_DATE"):
            vars_filled["LAST_SERVICE_DATE"] = case.get("expiry_dt", "")
        if "PHONE_NUMBER_FIRST4" in selected_vars and not vars_filled.get("PHONE_NUMBER_FIRST4"):
            mobile = case.get("customer_mobile", "")
            vars_filled["PHONE_NUMBER_FIRST4"] = mobile[:4] if len(mobile) >= 4 else mobile
        if "CALL_DATE" in selected_vars and not vars_filled.get("CALL_DATE"):
            vars_filled["CALL_DATE"] = case.get("call_timestamp", "")

        rendered = render_email(language, selected_vars, vars_filled)
        recipient = test_email if is_test else (case.get("partner_email") or "")

        if not recipient:
            results.append({"case_id": cid, "ok": False, "error": "No partner email"})
            continue

        send_result = send_email(recipient, rendered["subject"], rendered["body_text"], rendered["body_html"])

        b1db.log_breach1_email(
            case_id=cid,
            partner_name=case.get("partner_name", ""),
            partner_email=case.get("partner_email", ""),
            recipient_email=recipient,
            case_type=1,
            language=language,
            variables_json=json.dumps(vars_filled),
            subject=rendered["subject"],
            body_preview=rendered["body_text"][:500],
            is_test=is_test,
            status="sent" if send_result["ok"] else "failed",
            error=send_result.get("error"),
        )

        if send_result["ok"] and not is_test:
            b1db.mark_breach1_email_sent([cid], 1)

        results.append({"case_id": cid, **send_result})

    return jsonify({"results": results, "total": len(results),
                    "sent": sum(1 for r in results if r.get("ok"))})


@app.route("/api/breach1/email-log")
def breach1_email_log():
    return jsonify(b1db.get_breach1_email_log())


@app.route("/api/breach1/escalation-preview")
def breach1_escalation_preview():
    """Preview rows to append to escalation Google Sheet."""
    try:
        from google_sheets import get_existing_escalation_customers, get_all_partner_emails
        existing_mobiles = get_existing_escalation_customers()
        emails = get_all_partner_emails()
        cases = b1db.get_breach1_cases()

        rows = []
        skipped = 0
        for c in sorted(cases, key=lambda x: (x.get("partner_name", ""), x.get("customer_mobile", ""))):
            mobile = (c.get("customer_mobile") or "").strip()
            if mobile in existing_mobiles:
                skipped += 1
                continue

            email_info = emails.get((c.get("partner_name") or "").lower(), {})
            partner_id = email_info.get("partner_id", "")
            partner_email = c.get("partner_email") or email_info.get("email", "")

            detection_date = c.get("call_timestamp") or ""
            email_done = "Yes" if c.get("email_state") == "sent" else "No"
            email_date = (c.get("email_sent_at") or "")[:10] if c.get("email_sent_at") else ""

            row = [
                partner_id,
                c.get("partner_name", ""),
                mobile,
                "Breach Rule 1 System Fundamentals",
                detection_date,
                "6month churn logic",
                "",      # Penalty Amount
                "No",    # Penalty Done
                "",      # Penalty Done Date
                email_done,
                email_date,
                "No",    # WhatsApp
                "",      # WhatsApp Date
                "",      # Partner Mobile
                c.get("partner_name", ""),
                partner_email,
                "",      # Link
                "",      # Comments
            ]
            rows.append(row)

        return jsonify({"ok": True, "rows": rows, "skipped": skipped, "count": len(rows)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/breach1/escalation-push", methods=["POST"])
def breach1_escalation_push():
    """Append rows to FP1 Escalation Google Sheet."""
    try:
        from google_sheets import append_escalation_rows, get_existing_escalation_customers, get_all_partner_emails
        existing_mobiles = get_existing_escalation_customers()
        emails = get_all_partner_emails()
        cases = b1db.get_breach1_cases()

        rows = []
        for c in sorted(cases, key=lambda x: (x.get("partner_name", ""), x.get("customer_mobile", ""))):
            mobile = (c.get("customer_mobile") or "").strip()
            if mobile in existing_mobiles:
                continue

            email_info = emails.get((c.get("partner_name") or "").lower(), {})
            partner_id = email_info.get("partner_id", "")
            partner_email = c.get("partner_email") or email_info.get("email", "")

            detection_date = c.get("call_timestamp") or ""
            email_done = "Yes" if c.get("email_state") == "sent" else "No"
            email_date = (c.get("email_sent_at") or "")[:10] if c.get("email_sent_at") else ""

            row = [
                partner_id,
                c.get("partner_name", ""),
                mobile,
                "Breach Rule 1 System Fundamentals",
                detection_date,
                "6month churn logic",
                "",
                "No",
                "",
                email_done,
                email_date,
                "No",
                "",
                "",
                c.get("partner_name", ""),
                partner_email,
                "",
                "",
            ]
            rows.append(row)

        if not rows:
            return jsonify({"ok": True, "appended": 0, "message": "All cases already in sheet"})

        appended = append_escalation_rows(rows)
        return jsonify({"ok": True, "appended": len(rows), "message": f"{len(rows)} row(s) added"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/breach1/sources")
def breach1_sources():
    return jsonify(b1db.get_breach1_sources())


@app.route("/api/breach1/set-action", methods=["POST"])
def breach1_set_action():
    """Set action_type on selected case_ids. Supports 'warning', 'penalty', 'pending' (reset)."""
    body = request.json or {}
    case_ids = body.get("case_ids", [])
    action_type = body.get("action_type", "")
    if action_type not in ("warning", "penalty", "pending"):
        return jsonify({"error": "action_type must be 'warning', 'penalty', or 'pending'"}), 400
    if not case_ids:
        return jsonify({"error": "No case_ids provided"}), 400
    if action_type == "pending":
        b1db.reset_breach1_action(case_ids)
    else:
        b1db.set_breach1_action_type(case_ids, action_type)
    return jsonify({"ok": True, "count": len(case_ids), "action_type": action_type})


@app.route("/api/breach1/mark-email-sent", methods=["POST"])
def breach1_mark_email_sent():
    """Manually mark selected B1 cases as email sent (for emails sent outside the system)."""
    body = request.json or {}
    case_ids = body.get("case_ids", [])
    if not case_ids:
        return jsonify({"error": "No case_ids provided"}), 400
    b1db.mark_breach1_email_sent(case_ids, case_type=1)
    return jsonify({"ok": True, "count": len(case_ids)})


@app.route("/api/breach1/set-email", methods=["POST"])
def breach1_set_email():
    """Manually set partner email on B1 cases — by case_ids or by partner_name."""
    body = request.json or {}
    email = (body.get("email") or "").strip()
    if not email:
        return jsonify({"error": "No email provided"}), 400
    case_ids = body.get("case_ids", [])
    partner_name = body.get("partner_name", "").strip()
    if case_ids:
        b1db.set_breach1_partner_email(case_ids, email)
        return jsonify({"ok": True, "count": len(case_ids)})
    elif partner_name:
        count = b1db.set_breach1_partner_email_by_name(partner_name, email)
        return jsonify({"ok": True, "count": count, "partner_name": partner_name})
    return jsonify({"error": "Provide case_ids or partner_name"}), 400


@app.route("/api/breach1/penalty-xlsx", methods=["POST"])
def breach1_penalty_xlsx():
    """Generate penalty XLSX for B1 cases marked as penalty."""
    body = request.json or {}
    case_ids = body.get("case_ids", [])
    if case_ids:
        cases = [b1db.get_breach1_case(cid) for cid in case_ids]
        cases = [c for c in cases if c and c.get("action_type") == "penalty"]
    else:
        cases = b1db.get_breach1_penalty_cases()
    if not cases:
        return jsonify({"error": "No penalty cases found"}), 400
    xlsx_bytes = actions.generate_b1_penalty_xlsx(cases)
    # Mark cases as csv_generated
    b1db.mark_b1_penalty_csv_generated([c["id"] for c in cases])
    return Response(xlsx_bytes,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=fp1_penalty.xlsx"})


@app.route("/api/breach1/upload-penalty-status", methods=["POST"])
def breach1_upload_penalty_status():
    """Upload penalty result Excel for B1 cases, match by AccountId."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    rows = []
    fname = (f.filename or "").lower()
    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(f.read()), read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        wb.close()
    else:
        import csv, io
        try:
            text = f.read().decode("utf-8-sig")
        except Exception:
            return jsonify({"error": "Could not read file"}), 400
        rows = list(csv.DictReader(io.StringIO(text)))

    matched = []
    unmatched = []
    for row in rows:
        partner_id = str(row.get("AccountId") or row.get("Partner Id") or "").strip()
        if not partner_id:
            continue
        status = str(row.get("Process Status") or "").strip().lower()
        if status and status != "yes":
            unmatched.append({"partner_id": partner_id, "reason": row.get("Reason", "")})
            continue
        result = b1db.mark_b1_penalty_uploaded(partner_id)
        if result["matched"]:
            matched.append(result)
        else:
            unmatched.append({"partner_id": partner_id})
    return jsonify({"ok": True, "matched_count": len(matched), "unmatched_count": len(unmatched),
                    "matched": matched, "unmatched": unmatched})


@app.route("/api/breach1/penalty-template")
def breach1_penalty_template():
    from email_sender import get_fp1_penalty_template_info
    return jsonify(get_fp1_penalty_template_info())


@app.route("/api/breach1/preview-penalty-email", methods=["POST"])
def breach1_preview_penalty():
    from email_sender import render_fp1_penalty_email
    body = request.json or {}
    language = body.get("language", "both")
    selected_vars = body.get("selected_vars", [])
    values = body.get("values", {})
    try:
        result = render_fp1_penalty_email(language, selected_vars, values)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/breach1/send-penalty-email", methods=["POST"])
def breach1_send_penalty_email():
    from email_sender import render_fp1_penalty_email, send_email
    body = request.json or {}
    case_ids = body.get("case_ids", [])
    language = body.get("language", "both")
    selected_vars = body.get("selected_vars", [])
    values = body.get("values", {})
    test_email = body.get("test_email", "").strip()
    is_test = bool(test_email)

    results = []
    for cid in case_ids:
        case = b1db.get_breach1_case(cid)
        if not case:
            results.append({"case_id": cid, "ok": False, "error": "Case not found"})
            continue

        vars_filled = dict(values)
        if "LAST_SERVICE_DATE" in selected_vars and not vars_filled.get("LAST_SERVICE_DATE"):
            vars_filled["LAST_SERVICE_DATE"] = case.get("expiry_dt", "")
        if "PHONE_NUMBER_FIRST4" in selected_vars and not vars_filled.get("PHONE_NUMBER_FIRST4"):
            mobile = case.get("customer_mobile", "")
            vars_filled["PHONE_NUMBER_FIRST4"] = mobile[:4] if len(mobile) >= 4 else mobile

        rendered = render_fp1_penalty_email(language, selected_vars, vars_filled)
        recipient = test_email if is_test else (case.get("partner_email") or "")

        if not recipient:
            results.append({"case_id": cid, "ok": False, "error": "No partner email"})
            continue

        send_result = send_email(recipient, rendered["subject"], rendered["body_text"], rendered["body_html"])

        b1db.log_breach1_email(
            case_id=cid,
            partner_name=case.get("partner_name", ""),
            partner_email=case.get("partner_email", ""),
            recipient_email=recipient,
            case_type=1,
            language=language,
            variables_json=json.dumps(vars_filled),
            subject=rendered["subject"],
            body_preview=rendered["body_text"][:500],
            is_test=is_test,
            status="sent" if send_result["ok"] else "failed",
            error=send_result.get("error"),
        )

        if send_result["ok"] and not is_test:
            b1db.mark_b1_penalty_email_sent([cid])

        results.append({"case_id": cid, **send_result})

    return jsonify({"results": results, "total": len(results),
                    "sent": sum(1 for r in results if r.get("ok"))})


@app.route("/api/breach1/manual-report", methods=["POST"])
def breach1_manual_report():
    """Accept a free-text manual report and create a B1 case."""
    import re
    body = request.json or {}
    report_text = (body.get("report_text") or "").strip()
    reported_by = (body.get("reported_by") or "").strip()
    if not report_text:
        return jsonify({"error": "report_text is required"}), 400

    # Extract fields from free text
    mobiles = re.findall(r'\b[6-9]\d{9}\b', report_text)
    customer_mobile = mobiles[0] if mobiles else ""

    # Try to match partner name from partner email list
    try:
        from google_sheets import get_all_partner_emails
        partner_emails = get_all_partner_emails()
    except Exception:
        partner_emails = {}

    partner_name = ""
    partner_email = ""
    partner_id = ""
    text_lower = report_text.lower()
    for pname_lower, info in partner_emails.items():
        if pname_lower in text_lower:
            partner_name = info.get("name", pname_lower)
            partner_email = info.get("email", "")
            partner_id = info.get("partner_id", "")
            break

    data = {
        "customer_mobile": customer_mobile,
        "partner_name": partner_name,
        "partner_email": partner_email,
        "partner_id": partner_id,
        "report_text": report_text,
        "reported_by": reported_by,
        "source": "manual_report",
    }
    b1db.upsert_breach1_case(data)
    return jsonify({"ok": True, "extracted": {
        "customer_mobile": customer_mobile,
        "partner_name": partner_name or "(not detected)",
    }})


# ── Breach 4 (Router Misuse) ──────────────────────────────────────────────────

@app.route("/api/breach4/sync", methods=["POST"])
def breach4_sync():
    """Import FP4 cases from Google Sheet."""
    try:
        from google_sheets import fetch_fp4_cases, get_all_partner_emails
        cases = fetch_fp4_cases()
        emails = get_all_partner_emails()

        new_count = 0
        updated_count = 0
        for c in cases:
            partner_name = (c.get("Partner Name") or "").strip()
            partner_id = (c.get("Partner Id") or "").strip()
            email_info = emails.get(partner_name.lower(), {})
            data = {
                "partner_id": partner_id,
                "partner_name": partner_name,
                "customer_details": (c.get("Customer Details") or "").strip(),
                "principle_broken": (c.get("Principle Broken") or "").strip(),
                "device_id": (c.get("Device Id") or "").strip(),
                "date_reported": (c.get("Date Reported") or "").strip(),
                "reporting_channel": (c.get("Reporting Channel") or "").strip(),
                "penalty_amount": (c.get("Penalty Amount") or "").strip(),
                "penalty_done": (c.get("Penalty Done") or "").strip(),
                "penalty_done_date": (c.get("Penalty Done Date") or "").strip(),
                "partner_email_comms": (c.get("Partner Email Comms Done") or "").strip(),
                "email_date": (c.get("Email Date") or "").strip(),
                "whatsapp_comms": (c.get("Partner Text/Whatsapp Comms Done") or "").strip(),
                "whatsapp_date": (c.get("Whatsapp Date") or "").strip(),
                "partner_mobile": (c.get("Partner Mobile") or "").strip(),
                "link": (c.get("Link") or "").strip(),
                "comments": (c.get("Comments") or "").strip(),
                "partner_email": email_info.get("email", ""),
            }
            # Skip rows with no real data
            if not data["partner_id"] and not data["partner_name"] and not data["device_id"]:
                continue

            existing = db.get_breach4_cases(search=data["device_id"])
            existing = [e for e in existing if e["partner_id"] == data["partner_id"]
                        and e["customer_details"] == data["customer_details"]]
            if not existing:
                new_count += 1
            else:
                updated_count += 1
            db.upsert_breach4_case(data)

        return jsonify({"ok": True, "new_cases": new_count, "updated": updated_count, "total_fetched": len(cases)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/breach4/cases")
def breach4_cases():
    partner = request.args.get("partner", "all")
    email_state = request.args.get("email_state", "all")
    search = request.args.get("search", "").strip() or None
    return jsonify(db.get_breach4_cases(partner=partner, email_state=email_state, search=search))


@app.route("/api/breach4/cases/<int:case_id>")
def breach4_case(case_id):
    case = db.get_breach4_case(case_id)
    if not case:
        return jsonify({"error": "Not found"}), 404
    return jsonify(case)


@app.route("/api/breach4/summary")
def breach4_summary():
    return jsonify(db.get_breach4_summary())


@app.route("/api/breach4/partners")
def breach4_partners():
    return jsonify(db.get_breach4_partners())


@app.route("/api/breach4/template")
def breach4_template():
    from email_sender import get_fp4_template_info
    return jsonify(get_fp4_template_info())


@app.route("/api/breach4/preview-email", methods=["POST"])
def breach4_preview():
    from email_sender import render_fp4_email
    body = request.json or {}
    language = body.get("language", "both")
    selected_vars = body.get("selected_vars", [])
    values = body.get("values", {})
    try:
        result = render_fp4_email(language, selected_vars, values)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/breach4/send-email", methods=["POST"])
def breach4_send_email():
    from email_sender import render_fp4_email, send_email
    body = request.json or {}
    case_ids = body.get("case_ids", [])
    language = body.get("language", "both")
    selected_vars = body.get("selected_vars", [])
    values = body.get("values", {})
    test_email = body.get("test_email", "").strip()
    is_test = bool(test_email)

    results = []
    for cid in case_ids:
        case = db.get_breach4_case(cid)
        if not case:
            results.append({"case_id": cid, "ok": False, "error": "Case not found"})
            continue

        vars_filled = dict(values)
        if "ROUTER_ID" in selected_vars and not vars_filled.get("ROUTER_ID"):
            vars_filled["ROUTER_ID"] = case.get("device_id", "")

        rendered = render_fp4_email(language, selected_vars, vars_filled)
        recipient = test_email if is_test else (case.get("partner_email") or "")

        if not recipient:
            results.append({"case_id": cid, "ok": False, "error": "No partner email"})
            continue

        send_result = send_email(recipient, rendered["subject"], rendered["body_text"], rendered["body_html"])

        db.log_breach4_email(
            case_id=cid,
            partner_name=case.get("partner_name", ""),
            partner_email=case.get("partner_email", ""),
            recipient_email=recipient,
            language=language,
            variables_json=json.dumps(vars_filled),
            subject=rendered["subject"],
            body_preview=rendered["body_text"][:500],
            is_test=is_test,
            status="sent" if send_result["ok"] else "failed",
            error=send_result.get("error"),
        )

        if send_result["ok"] and not is_test:
            db.mark_breach4_email_sent([cid])

        results.append({"case_id": cid, **send_result})

    return jsonify({"results": results, "total": len(results),
                    "sent": sum(1 for r in results if r.get("ok"))})


@app.route("/api/breach4/email-log")
def breach4_email_log():
    return jsonify(db.get_breach4_email_log())


@app.route("/api/breach4/escalation-preview")
def breach4_escalation_preview():
    """Preview rows to update in FP4 Google Sheet."""
    try:
        from google_sheets import get_existing_fp4_customers
        existing = get_existing_fp4_customers()
        cases = db.get_breach4_cases()

        rows = []
        skipped = 0
        for c in sorted(cases, key=lambda x: (x.get("partner_name", ""), x.get("customer_details", ""))):
            cust = (c.get("customer_details") or "").strip()
            email_done = "Yes" if c.get("email_state") == "sent" else c.get("partner_email_comms", "No")
            email_date = (c.get("email_sent_at") or "")[:10] if c.get("email_sent_at") else c.get("email_date", "")

            row = [
                c.get("partner_id", ""),
                c.get("partner_name", ""),
                cust,
                c.get("principle_broken", ""),
                c.get("device_id", ""),
                c.get("date_reported", ""),
                c.get("reporting_channel", ""),
                c.get("penalty_amount", ""),
                c.get("penalty_done", ""),
                c.get("penalty_done_date", ""),
                email_done,
                email_date,
                c.get("whatsapp_comms", ""),
                c.get("whatsapp_date", ""),
                c.get("partner_mobile", ""),
                c.get("link", ""),
                c.get("comments", ""),
            ]
            rows.append(row)

        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/breach4/escalation-push", methods=["POST"])
def breach4_escalation_push():
    """Overwrite FP4 sheet data rows with current local data."""
    try:
        from google_sheets import _get_service, ESCALATION_SHEET_ID
        cases = db.get_breach4_cases()

        rows = []
        for c in sorted(cases, key=lambda x: (x.get("partner_name", ""), x.get("customer_details", ""))):
            email_done = "Yes" if c.get("email_state") == "sent" else c.get("partner_email_comms", "No")
            email_date = (c.get("email_sent_at") or "")[:10] if c.get("email_sent_at") else c.get("email_date", "")

            row = [
                c.get("partner_id", ""),
                c.get("partner_name", ""),
                c.get("customer_details", ""),
                c.get("principle_broken", ""),
                c.get("device_id", ""),
                c.get("date_reported", ""),
                c.get("reporting_channel", ""),
                c.get("penalty_amount", ""),
                c.get("penalty_done", ""),
                c.get("penalty_done_date", ""),
                email_done,
                email_date,
                c.get("whatsapp_comms", ""),
                c.get("whatsapp_date", ""),
                c.get("partner_mobile", ""),
                c.get("link", ""),
                c.get("comments", ""),
            ]
            rows.append(row)

        if not rows:
            return jsonify({"ok": True, "appended": 0, "message": "No cases to push"})

        service = _get_service(readonly=False)
        # Clear existing data rows (keep header)
        service.spreadsheets().values().clear(
            spreadsheetId=ESCALATION_SHEET_ID,
            range="'FP4 : Router Misuse'!A2:Q5000",
        ).execute()
        # Write updated rows
        body = {"values": rows}
        service.spreadsheets().values().update(
            spreadsheetId=ESCALATION_SHEET_ID,
            range="'FP4 : Router Misuse'!A2",
            valueInputOption="USER_ENTERED",
            body=body,
        ).execute()

        return jsonify({"ok": True, "appended": len(rows), "message": f"{len(rows)} row(s) updated"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Settings ──────────────────────────────────────────────────────────────────

@app.route("/api/settings")
def get_settings():
    cfg  = config.load()
    safe = dict(cfg)
    if safe.get("metabase_password"):
        safe["metabase_password"] = "••••••••"
    return jsonify(safe)


@app.route("/api/settings", methods=["POST"])
def save_settings():
    body = request.json or {}
    if body.get("metabase_password") == "••••••••":
        body.pop("metabase_password")
    config.save(body)
    return jsonify({"ok": True})


@app.route("/api/settings/test-metabase", methods=["POST"])
def test_metabase():
    cfg = config.load()
    try:
        from metabase import MetabaseClient
        client = MetabaseClient(
            cfg["metabase_url"], cfg.get("metabase_database_id", ""),
            api_key=cfg.get("metabase_api_key", ""),
            username=cfg.get("metabase_username", ""),
            password=cfg.get("metabase_password", ""),
        )
        ok = client.test_connection()
        return jsonify({"ok": ok})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/settings/databases", methods=["GET"])
def list_databases():
    cfg = config.load()
    try:
        from metabase import MetabaseClient
        client = MetabaseClient(
            cfg["metabase_url"], cfg.get("metabase_database_id", ""),
            api_key=cfg.get("metabase_api_key", ""),
            username=cfg.get("metabase_username", ""),
            password=cfg.get("metabase_password", ""),
        )
        dbs = client.list_databases()
        return jsonify({"ok": True, "databases": dbs})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/settings/test-slack", methods=["POST"])
def test_slack():
    cfg    = config.load()
    result = actions.send_to_slack(
        cfg.get("slack_webhook_url", ""),
        ":white_check_mark: WIOM Breach Tracker — Slack test successful!",
    )
    return jsonify(result)


if __name__ == "__main__":
    app.run(debug=False, port=5050, use_reloader=False, threaded=True)
